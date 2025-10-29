import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter

# Default ref locations (override when calling functions)
DEFAULT_REFS = {
    "category": r"C:\Users\User\OneDrive\Kimberlin Enterprises\REPORTS\References\CATEGORY.xlsx",
    "field_supervisors": r"C:\Users\User\OneDrive\Kimberlin Enterprises\REPORTS\References\Field Supervisors.xlsx",
    "week": r"C:\Users\User\OneDrive\Kimberlin Enterprises\REPORTS\References\WEEK.xlsx",
}

DEFAULT_INPUT_FILES = {
    "invoice": "DMS-Invoice-on.xlsx",
    "returns": "DMS-Customer Returns-on.xlsx",
    "customer": "DMS-Customer-on.xlsx",
    "pricelist": "DMS-Price-on.xlsx",
    "sales_order": "DMS-Sales Order-on.xlsx",
}

ACCOUNTING_FMT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)'
ALLOWED_PRODUCT_COLS = ['Product Code', 'product_code', 'SKU CODE']

def read_refs(ref_paths=None):
    ref_paths = ref_paths or DEFAULT_REFS
    return {k: pd.read_excel(v) for k, v in ref_paths.items()}

def read_inputs(import_path, input_files=None):
    files = input_files or DEFAULT_INPUT_FILES
    return {k: pd.read_excel(os.path.join(import_path, fname)) for k, fname in files.items()}

def build_m0_pricelist(pl_df):
    df = pl_df[['product_code', 'product_description', 'uom_description', 'selling_price', 'cust_class', 'cust_channel']].copy()
    df['with_vat'] = df['selling_price'] * 1.12
    df = df[(df['cust_class'] != 'BEV Dealer') & (df['cust_channel'] != 'VAN(EXTRUCK)')]
    pivot = df.pivot_table(index=['product_code', 'product_description'], columns='uom_description', values='with_vat', aggfunc='sum').reset_index()
    for col in ('Case','Subcase','Piece'):
        if col not in pivot.columns:
            pivot[col] = 0
    pl_m0 = pivot[['product_code','product_description','Case','Subcase','Piece']].copy()

    # canonical price_ref contains both SKU CODE and Product Code names to avoid merge KeyErrors
    price_ref = pl_m0[['product_code','Case']].rename(columns={'product_code':'SKU CODE','Case':'SKU PRICE REFERENCE'}).copy()
    price_ref['Product Code'] = price_ref['SKU CODE']
    return pl_m0, price_ref

def _ensure_price_ref(price_ref):
    """Return DataFrame with columns ['Product Code','SKU PRICE REFERENCE'].
    Only recognise product col names in ALLOWED_PRODUCT_COLS.
    """
    if price_ref is None:
        return pd.DataFrame(columns=['Product Code','SKU PRICE REFERENCE'])

    # normalize into DataFrame
    if isinstance(price_ref, pd.Series):
        pr = price_ref.reset_index()
    else:
        pr = price_ref.copy()

    if not isinstance(pr, pd.DataFrame):
        pr = pd.DataFrame(pr)

    # rename any allowed product col to canonical 'Product Code'
    for c in ALLOWED_PRODUCT_COLS:
        if c in pr.columns:
            if c != 'Product Code':
                pr = pr.rename(columns={c: 'Product Code'})
            break

    # choose price column: prefer 'SKU PRICE REFERENCE' or 'Case', else second column
    if 'SKU PRICE REFERENCE' not in pr.columns and 'Case' in pr.columns:
        pr = pr.rename(columns={'Case': 'SKU PRICE REFERENCE'})

    if 'SKU PRICE REFERENCE' not in pr.columns:
        # if there's more than one column, take the first non-product column as price
        candidates = [col for col in pr.columns if col != 'Product Code']
        if candidates:
            pr = pr.rename(columns={candidates[0]: 'SKU PRICE REFERENCE'})
        else:
            pr['SKU PRICE REFERENCE'] = 0

    # ensure canonical columns exist
    if 'Product Code' not in pr.columns:
        pr['Product Code'] = None
    if 'SKU PRICE REFERENCE' not in pr.columns:
        pr['SKU PRICE REFERENCE' ] = 0

    # cast types
    pr['Product Code'] = pr['Product Code'].astype(object)
    pr['SKU PRICE REFERENCE'] = pd.to_numeric(pr['SKU PRICE REFERENCE'], errors='coerce').fillna(0)

    return pr[['Product Code','SKU PRICE REFERENCE']].copy()

def format_and_save_excel(df, out_path, numeric_cols=None, autofit_cols=None, header_row=2):
    numeric_cols = numeric_cols or []
    autofit_cols = autofit_cols or []
    tmp = os.path.join(os.path.dirname(out_path) or '.', "_tmp_report_tools.xlsx")
    with pd.ExcelWriter(tmp, engine='openpyxl') as w:
        df.to_excel(w, index=False, startrow=header_row-1)
    wb = load_workbook(tmp)
    ws = wb.active
    # header style
    hf = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    hf_font = Font(bold=True, color='FFFFFF')
    for cell in ws[header_row]:
        cell.fill = hf
        cell.font = hf_font
        cell.alignment = Alignment(horizontal='center')
    # robust named styles check
    existing_names = set(getattr(s, 'name', s) for s in wb.named_styles)
    if "accounting_style" not in existing_names:
        ns = NamedStyle(name="accounting_style", number_format=ACCOUNTING_FMT)
        wb.add_named_style(ns)
    data_start = header_row + 1
    data_end = ws.max_row
    sum_row = data_end + 1

    def col_to_index(col):
        if isinstance(col, int):
            return col
        try:
            return list(df.columns).index(col) + 1
        except ValueError:
            return None

    for col in numeric_cols:
        idx = col_to_index(col)
        if not idx:
            continue
        for r in range(data_start, data_end + 1):
            cell = ws.cell(row=r, column=idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = ACCOUNTING_FMT
        col_letter = get_column_letter(idx)
        s = ws.cell(row=sum_row, column=idx)
        s.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        s.font = Font(bold=True, color="FF0000")
        s.number_format = ACCOUNTING_FMT

    for col in autofit_cols:
        idx = col_to_index(col) or (col if isinstance(col,int) else None)
        if not idx:
            continue
        letter = get_column_letter(idx) if isinstance(idx,int) else idx
        maxlen = 0
        for c in ws[letter]:
            if c.value:
                maxlen = max(maxlen, len(str(c.value)))
        ws.column_dimensions[letter].width = maxlen + 2

    ws.sheet_view.showGridLines = False
    wb.save(out_path)
    wb.close()
    try:
        os.remove(tmp)
    except OSError:
        pass

def _build_price_map(price_ref):
    """Map string(product_code) -> numeric price using only allowed product headers."""
    pr = _ensure_price_ref(price_ref)
    # keep keys as strings for reliable mapping
    keys = pr['Product Code'].astype(str)
    vals = pd.to_numeric(pr['SKU PRICE REFERENCE'], errors='coerce').fillna(0)
    return dict(zip(keys, vals))

def assemble_net_invoiced(inp, refs, price_ref):
    inv = inp['invoice'].copy()
    inv = inv[inv['Invoice Item Type'] != 'ITM_SALES_TAX'][[
        'Invoice Date','Sold To Customer Number','Sold To Customer Name','Product Code','Product/Item Description','Total Item amount with Tax and Discount'
    ]]
    inv.columns = ['DATE','ACCOUNT CODE','ACCOUNT NAME','SKU CODE','SKU NAME','SERVED INVOICE']
    inv['SERVED INVOICE'] = inv['SERVED INVOICE'].fillna(0)
    inv['BAD RETURNS'] = 0; inv['GOOD RETURNS'] = 0

    ret = inp['returns'][[
        'Customer Return Date','Sold To Customer Number','Sold To Customer Name','Product Code','Product Description','Facility Name','Estimated Product Return Amount'
    ]].copy()
    ret['Estimated Product Return Amount'] = ret['Estimated Product Return Amount'].fillna(0)
    ret['with_vat'] = ret['Estimated Product Return Amount'] * 1.12
    cust_ret = ret.pivot_table(index=['Customer Return Date','Sold To Customer Number','Sold To Customer Name','Product Code','Product Description'], columns='Facility Name', values='with_vat', aggfunc='sum').reset_index()
    for c in ('BO','FG'):
        if c not in cust_ret.columns:
            cust_ret[c]=0
    cust_ret = cust_ret.rename(columns={
        'Customer Return Date':'DATE','Sold To Customer Number':'ACCOUNT CODE','Sold To Customer Name':'ACCOUNT NAME',
        'Product Code':'SKU CODE','Product Description':'SKU NAME','BO':'BAD RETURNS','FG':'GOOD RETURNS'
    })
    cust_ret['SERVED INVOICE']=0
    net = pd.concat([inv, cust_ret], ignore_index=True, sort=False)

    cust_df = inp['customer'][[
        'NEXT_UP_NUMBER','PARTY_CLASSIFICATION_DESCRIPTION','KEY_ACCOUNT','SALES_REP_ID','SALES_REP_NAME','BRANCH_NAME','GEO_LOCATION_HIERARCHYDESCRIPTION','CITY','STATE_PROVINCE','CHANNEL'
    ]].rename(columns={'NEXT_UP_NUMBER':'ACCOUNT CODE'})
    net = net.merge(cust_df, on='ACCOUNT CODE', how='left')
    net = net.merge(refs['category'].drop(columns=['SKU NAME'], errors='ignore'), on='SKU CODE', how='left')
    net = net.merge(refs['field_supervisors'], on='SALES_REP_ID', how='left')
    net = net.merge(refs['week'], on='DATE', how='inner')

    # previous: fragile merge -> now map
    price_map = _build_price_map(price_ref)
    # map using SKU CODE column from net (converted to string)
    net['SKU PRICE REFERENCE'] = net['SKU CODE'].astype(str).map(price_map).fillna(0)

    # prefer SKU PRICE REFERENCE column
    net['SKU PRICE REFERENCE'] = pd.to_numeric(net.get('SKU PRICE REFERENCE', 0), errors='coerce').fillna(0)
    net['VALUE'] = net['SERVED INVOICE'].fillna(0) - net['BAD RETURNS'].fillna(0) - net['GOOD RETURNS'].fillna(0)
    net['VOLUME'] = 0
    mask = net['SKU PRICE REFERENCE'] != 0
    net.loc[mask,'VOLUME'] = net.loc[mask,'VALUE'] / net.loc[mask,'SKU PRICE REFERENCE']

    net['RD Name'] = 'Kimberlin'
    rename_map = {
        'DATE':'Date','WEEK':'Week','BRANCH_NAME':'Branch Name','SALES_REP_ID':'Employee Code','SALES_REP_NAME':'Employee Name',
        'KEY_ACCOUNT':'Channel','ACCOUNT CODE':'Sold To Customer number','ACCOUNT NAME':'Sold To Customer Name','CATEGORY':'Category',
        'SKU CODE':'Product Code','SKU NAME':'Product Description','VOLUME':'Volume','VALUE':'Net Value','GOOD RETURNS':'Good Stock Returns',
        'BAD RETURNS':'Bad Stock Returns','PARTY_CLASSIFICATION_DESCRIPTION':'Channel_Classification','GEO_LOCATION_HIERARCHYDESCRIPTION':'Brgy',
        'CITY':'Town','STATE_PROVINCE':'Province','FS':'FS','CHANNEL':'RTM Model'
    }
    net.rename(columns=rename_map, inplace=True)
    cols = ['RD Name','Date','Week','Branch Name','Employee Code','Employee Name','Channel','Sold To Customer number','Sold To Customer Name','Category',
            'Product Code','Product Description','Volume','Net Value','Good Stock Returns','Bad Stock Returns','Channel_Classification','Brgy','Town','Province','FS','RTM Model']
    return net[cols].sort_values(by=['Date','Sold To Customer number'])

def assemble_served_invoice(inp, refs, price_ref):
    inv = inp['invoice'][inp['invoice']['Invoice Item Type']!='ITM_SALES_TAX'][[
        'Invoice Date','Sold To Customer Number','Sold To Customer Name','Product Code','Product/Item Description','Total Item amount with Tax and Discount'
    ]].copy()
    inv.columns = ['DATE','ACCOUNT CODE','ACCOUNT NAME','SKU CODE','SKU NAME','SERVED INVOICE']
    inv['SERVED INVOICE'] = inv['SERVED INVOICE'].fillna(0)

    cust_df = inp['customer'][['NEXT_UP_NUMBER','PARTY_CLASSIFICATION_DESCRIPTION','KEY_ACCOUNT','SALES_REP_ID','SALES_REP_NAME','BRANCH_NAME','GEO_LOCATION_HIERARCHYDESCRIPTION','CITY','STATE_PROVINCE','CHANNEL']].rename(columns={'NEXT_UP_NUMBER':'ACCOUNT CODE'})
    df = (inv.merge(cust_df, on='ACCOUNT CODE', how='left')
             .merge(refs['category'].drop(columns=['SKU NAME'],errors='ignore'), on='SKU CODE', how='left')
             .merge(refs['field_supervisors'], on='SALES_REP_ID', how='left')
             .merge(refs['week'], on='DATE', how='inner')
         )

    price_map = _build_price_map(price_ref)
    df['SKU PRICE REFERENCE'] = df['SKU CODE'].astype(str).map(price_map).fillna(0)
    df['VOLUME'] = 0
    mask = df['SKU PRICE REFERENCE'] != 0
    df.loc[mask,'VOLUME'] = df.loc[mask,'SERVED INVOICE'] / df.loc[mask,'SKU PRICE REFERENCE']
    df['RD Name']='Kimberlin'
    df.rename(columns={'DATE':'Invoice Date','WEEK':'Week','BRANCH_NAME':'Branch Name','SALES_REP_ID':'Employee Code','SALES_REP_NAME':'Employee Name','KEY_ACCOUNT':'Channel','ACCOUNT CODE':'Sold To Customer Number','ACCOUNT NAME':'Sold To Customer Name','CATEGORY':'Category','SKU CODE':'Product Code','SKU NAME':'Product Description','VOLUME':'Volume','SERVED INVOICE':'Value','PARTY_CLASSIFICATION_DESCRIPTION':'Channel Type'}, inplace=True)
    cols = ['RD Name','Invoice Date','Week','Branch Name','Employee Code','Employee Name','Channel','Sold To Customer Number','Sold To Customer Name','Category','Product Code','Product Description','Volume','Value','FS','Channel Type']
    return df[cols].sort_values(by=['Invoice Date','Sold To Customer Number'])

def assemble_sales_orders(inp, refs, price_ref):
    """Assemble sales orders report with volume calculations."""
    # 1. Extract and prepare sales orders
    so = inp['sales_order'][[
        'Last Modified Date',
        'Sold To Customer number',
        'Sold To Customer Name',
        'Product Code',
        'Product Description',
        'Total Product Amount',
        'SO status'
    ]].copy()
    
    # Filter and calculate VAT
    so = so[so['SO status'] == 'Invoiced']
    so['Total Product Amount'] = pd.to_numeric(so['Total Product Amount'], errors='coerce').fillna(0)
    so['with vat'] = so['Total Product Amount'] * 1.12

    # 2. Prepare customer data
    cust = inp['customer'][[
        'NEXT_UP_NUMBER',
        'PARTY_CLASSIFICATION_DESCRIPTION',
        'KEY_ACCOUNT',
        'SALES_REP_ID',
        'SALES_REP_NAME',
        'BRANCH_NAME',
        'GEO_LOCATION_HIERARCHYDESCRIPTION',
        'CITY',
        'STATE_PROVINCE',
        'CHANNEL'
    ]].rename(columns={'NEXT_UP_NUMBER': 'Sold To Customer number'})

    # 3. Build main dataframe through merges
    df = (so.merge(cust, on='Sold To Customer number', how='left')
            .merge(refs['category'].rename(columns={'SKU CODE': 'Product Code'}), 
                  on='Product Code', how='left')
            .merge(refs['field_supervisors'], on='SALES_REP_ID', how='left')
            .merge(refs['week'].rename(columns={'DATE': 'Last Modified Date'}), 
                  on='Last Modified Date', how='inner')
         )

    # 4. Price reference and volume calculation
    pr = _ensure_price_ref(price_ref)
    df['SKU PRICE REFERENCE'] = pd.to_numeric(pr['SKU PRICE REFERENCE'], errors='coerce').fillna(0)
    
    # Initialize VOLUME as float64 to avoid dtype warning
    df['VOLUME'] = pd.Series(0, index=df.index, dtype='float64')
    mask = df['SKU PRICE REFERENCE'] != 0
    if mask.any():
        df.loc[mask, 'VOLUME'] = (
            pd.to_numeric(df.loc[mask, 'with vat'], errors='coerce') / 
            pd.to_numeric(df.loc[mask, 'SKU PRICE REFERENCE'], errors='coerce')
        ).fillna(0)

    # 5. Add RD name and standardize columns
    df['RD Name'] = 'Kimberlin'
    df.rename(columns={
        'Last Modified Date': 'SO Date',
        'WEEK': 'Week',
        'BRANCH_NAME': 'Branch Name',
        'SALES_REP_ID': 'Employee Code',
        'SALES_REP_NAME': 'Employee Name',
        'KEY_ACCOUNT': 'Channel',
        'Sold To Customer number': 'Sold To Customer Number',
        'CATEGORY': 'Category',
        'VOLUME': 'Volume',
        'with vat': 'Value',
        'PARTY_CLASSIFICATION_DESCRIPTION': 'Channel Type'
    }, inplace=True)

    # 6. Select and order final columns
    cols = [
        'RD Name', 'SO Date', 'Week', 'Branch Name', 'Employee Code',
        'Employee Name', 'Channel', 'Sold To Customer Number',
        'Sold To Customer Name', 'Category', 'Product Code',
        'Product Description', 'Volume', 'Value', 'FS', 'Channel Type'
    ]
    
    return df[cols].sort_values(by=['SO Date', 'Sold To Customer Number'])